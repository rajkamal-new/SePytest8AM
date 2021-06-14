from openpyxl import load_workbook


class Excel:

    def __init__(self, sheetname, filename = r"C:\Users\rajka\NewPycharmProjects\SePytest8AM\data\Test_Data.xlsx"):
        self.wb = load_workbook(filename)
        self.sheet = self.wb[sheetname]

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, exc_traceback):
        self.wb.close()

    def get_max_row_col(self):
        self.max_row =  self.sheet.max_row+1
        self.max_col = self.sheet.max_column+1

    def get_cell_data(self, row, col):
        value = self.sheet.cell(row=row, column=col).value
        if value is None:
            return ""
        return value

    def get_data(self):
        _data = []
        self.get_max_row_col()
        for row in range(2, self.max_row):
            if self.get_cell_data(row, 2) != "y":
                continue
            _row_data = []
            for col in range(3, self.max_col):
                _row_data.append(self.get_cell_data(row, col))

            _data.append(_row_data)

        return _data

    def close(self):
        self.wb.close()

if __name__ == '__main__':

    xl = Excel("LoginFail")
    xl.get_data()
    xl.close()



